from flask import Flask, render_template, request, jsonify, send_file
import google.generativeai as genai
import os
import json
from datetime import datetime, timedelta
import sqlite3
from pathlib import Path
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv

# .env 파일 로드
load_dotenv()

app = Flask(__name__)

# Gemini API 설정
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')
if not GEMINI_API_KEY:
    print("⚠️  경고: GEMINI_API_KEY가 설정되지 않았습니다!")
    print("   .env 파일에 GEMINI_API_KEY=your-key를 추가하세요.")
else:
    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel('gemini-2.0-flash-exp')

# 서비스 목록 - load_policies()에서 동적으로 감지됨
SERVICES = []

# 정책 데이터 저장
SERVICE_POLICIES = {}

# DB 초기화
def init_db():
    conn = sqlite3.connect('quiz_results.db')
    c = conn.cursor()
    
    # 결과 테이블
    c.execute('''
        CREATE TABLE IF NOT EXISTS results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            user_id TEXT,
            name TEXT,
            center TEXT,
            service TEXT,
            role TEXT,
            score INTEGER,
            total INTEGER,
            time_spent INTEGER,
            timeout INTEGER DEFAULT 0,
            questions_data TEXT,
            answers_data TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 관리자 프롬프트 테이블
    c.execute('''
        CREATE TABLE IF NOT EXISTS admin_prompts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            month TEXT,
            service TEXT,
            custom_prompt TEXT,
            difficulty TEXT,
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            is_active INTEGER DEFAULT 1
        )
    ''')
    
    conn.commit()
    conn.close()
    print("✅ SQLite 데이터베이스 초기화 완료")

# 정책 파일 로딩
def load_policies():
    global SERVICES, SERVICE_POLICIES
    
    policies_dir = Path('policies')
    if not policies_dir.exists():
        print("⚠️  정책 파일이 없습니다!")
        print("   policies/ 폴더를 생성하고 정책 파일을 추가하세요.")
        return
    
    print("\n" + "="*60)
    print("📚 정책 파일 로딩 중...")
    print("="*60)
    
    # 서비스 자동 감지
    detected_services = set()
    
    # 방식 1: 폴더별 서비스 감지
    for item in policies_dir.iterdir():
        if item.is_dir():
            detected_services.add(item.name)
    
    # 방식 2: 파일명에서 서비스 추출 (service_정책.md 형태)
    for policy_file in policies_dir.glob('*.md'):
        filename = policy_file.stem
        if '_' in filename:
            service = filename.split('_')[0]
            detected_services.add(service)
    
    # SERVICES 업데이트
    SERVICES = sorted(list(detected_services))
    SERVICE_POLICIES = {}
    
    print(f"🔍 감지된 서비스: {', '.join(SERVICES)}")
    print(f"📊 총 {len(SERVICES)}개 서비스")
    print("-" * 60)
    
    loaded_count = 0
    
    # 방식 1: 폴더별 정책 로딩
    for service in SERVICES:
        service_dir = policies_dir / service
        if service_dir.exists() and service_dir.is_dir():
            all_policies = []
            for policy_file in service_dir.glob('*.md'):
                try:
                    with open(policy_file, 'r', encoding='utf-8') as f:
                        content = f.read()
                        all_policies.append(content)
                        loaded_count += 1
                        print(f"✅ 로드 완료: {service}/{policy_file.name}")
                except Exception as e:
                    print(f"❌ 로드 실패: {service}/{policy_file.name} - {e}")
            
            if all_policies:
                SERVICE_POLICIES[service] = "\n\n".join(all_policies)
    
    # 방식 2: 파일명 기반 로딩 (택시_정책.md)
    for policy_file in policies_dir.glob('*.md'):
        try:
            filename = policy_file.stem
            if '_' in filename:
                service = filename.split('_')[0]
                if service in SERVICES:
                    with open(policy_file, 'r', encoding='utf-8') as f:
                        content = f.read()
                        if service in SERVICE_POLICIES:
                            SERVICE_POLICIES[service] += "\n\n" + content
                        else:
                            SERVICE_POLICIES[service] = content
                        loaded_count += 1
                        print(f"✅ 로드 완료: {policy_file.name}")
        except Exception as e:
            print(f"❌ 로드 실패: {policy_file.name} - {e}")
    
    print("\n📚 총 {}개 서비스의 정책이 로드되었습니다.".format(len(SERVICE_POLICIES)))
    for service, policy in SERVICE_POLICIES.items():
        print(f"   - {service}: {len(policy):,}자")
    print("="*60 + "\n")

# 재응시 체크 제거됨 (언제든지 응시 가능)

# 관리자 프롬프트 가져오기
def get_admin_prompt(service):
    conn = sqlite3.connect('quiz_results.db')
    c = conn.cursor()
    
    current_month = datetime.now().strftime('%Y-%m')
    c.execute('''
        SELECT custom_prompt, difficulty FROM admin_prompts 
        WHERE month = ? AND service = ? AND is_active = 1
        ORDER BY created_at DESC LIMIT 1
    ''', (current_month, service))
    
    result = c.fetchone()
    conn.close()
    
    if result:
        return {'custom_prompt': result[0], 'difficulty': result[1]}
    return None

@app.route('/')
def index():
    return send_file('index_new.html')

@app.route('/results')
def results():
    return send_file('results.html')

@app.route('/admin')
def admin():
    return send_file('admin.html')

@app.route('/api/services')
def get_services():
    return jsonify({
        'success': True,
        'services': list(SERVICE_POLICIES.keys())
    })

@app.route('/generate-quiz', methods=['POST'])
def generate_quiz_alt():
    # index_new.html에서 호출하는 경로
    return generate_quiz()

@app.route('/api/generate-quiz', methods=['POST'])
def generate_quiz():
    try:
        data = request.json
        service = data.get('service')
        role = data.get('role', '상담사')
        num_questions = 10
        
        if service not in SERVICE_POLICIES:
            return jsonify({
                'success': False,
                'error': f'{service} 서비스의 정책이 로드되지 않았습니다.'
            }), 400
        
        # 선택 서비스 정책 (70%)
        selected_policy = SERVICE_POLICIES[service]
        
        # T공통 정책 (30%)
        common_policy = SERVICE_POLICIES.get('T공통', '')
        
        # 관리자 커스텀 프롬프트 가져오기
        admin_config = get_admin_prompt(service)
        
        # 역할별 난이도 설정 (다양한 역할 지원)
        difficulty_map = {
            '신입상담사': '초급',
            '상담사(신입)': '초급',
            '3개월 미만 상담사': '초급',
            '상담사': '중급',
            '경력상담사': '중급',
            '3개월 이상 상담사': '중급',
            '관리자': '고급',
            '팀장': '고급',
            '매니저': '고급',
            'TL': '고급'
        }
        difficulty = difficulty_map.get(role, '중급')
        
        # 관리자 커스텀 난이도가 있으면 우선
        if admin_config:
            difficulty = admin_config.get('difficulty', difficulty)
        
        # 프롬프트 구성
        prompt = f"""당신은 카카오 T 고객센터 직무능력 테스트 출제 전문가입니다.

## 출제 요구사항

**대상**: {role}
**서비스**: {service}
**난이도**: {difficulty}
**문제 구성**: 총 10문제 (객관식 8문제, 주관식 2문제)
**정책 비율**: 
- {service} 서비스 정책: 70% (7문제)
- T공통 정책: 30% (3문제)

## {service} 서비스 정책:
{selected_policy[:15000]}

## T공통 정책:
{common_policy[:5000]}

"""

        # 관리자 커스텀 프롬프트 추가
        if admin_config and admin_config.get('custom_prompt'):
            prompt += f"\n## 관리자 추가 요구사항:\n{admin_config['custom_prompt']}\n"
        
        prompt += f"""
## 난이도별 출제 기준:

### 초급 ({role}이(가) 신입인 경우):
- 기본 용어와 개념
- 단순한 절차와 프로세스
- 자주 사용하는 기능
- 명확한 답이 있는 문제

### 중급 ({role}이(가) 일반인 경우):
- 실무 상황 대처
- 정책 적용 방법
- 고객 응대 시나리오
- 예외 상황 처리

### 고급 ({role}이(가) 관리자인 경우):
- 복합적인 상황 판단
- 정책 이해도 심화
- 의사결정 능력
- 문제 해결 능력

## 출력 형식 (반드시 JSON만):

{{
  "questions": [
    {{
      "id": 1,
      "type": "객관식",
      "question": "질문 내용",
      "options": ["선택지1", "선택지2", "선택지3", "선택지4"],
      "answer": "정답",
      "explanation": "상세한 설명"
    }},
    {{
      "id": 2,
      "type": "주관식",
      "question": "질문 내용",
      "answer": "모범 답안",
      "explanation": "상세한 설명"
    }}
  ]
}}

**중요**: 반드시 유효한 JSON만 출력하고, 다른 텍스트는 포함하지 마세요.
"""
        
        print(f"\n🤖 AI 퀴즈 생성 중...")
        print(f"   서비스: {service}")
        print(f"   역할: {role}")
        print(f"   난이도: {difficulty}")
        if admin_config:
            print(f"   ✨ 관리자 커스텀 프롬프트 적용됨")
        
        response = model.generate_content(prompt)
        response_text = response.text.strip()
        
        # JSON 추출
        if '```json' in response_text:
            response_text = response_text.split('```json')[1].split('```')[0].strip()
        elif '```' in response_text:
            response_text = response_text.split('```')[1].split('```')[0].strip()
        
        quiz_data = json.loads(response_text)
        
        print(f"✅ 퀴즈 생성 완료!")
        
        return jsonify({
            'success': True,
            'questions': quiz_data['questions']
        })
        
    except Exception as e:
        print(f"❌ 오류 발생: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/submit-exam', methods=['POST'])
def submit_exam_alt():
    # index_new.html에서 호출하는 경로
    return submit_quiz()

@app.route('/api/submit', methods=['POST'])
def submit_quiz():
    try:
        data = request.json
        
        # 채점
        questions = data.get('questions', [])
        answers = data.get('answers', [])
        
        score = 0
        total = len(questions)
        
        for i, question in enumerate(questions):
            user_answer = answers[i] if i < len(answers) else None
            
            if question.get('type') == '객관식':
                # 객관식: 정확히 일치해야 정답
                correct_answer = question.get('answer')
                if user_answer == correct_answer:
                    score += 1
            else:
                # 주관식: 답변이 있으면 1점 (간단한 평가)
                # 실제로는 더 정교한 평가가 필요하지만, 일단 기본 구현
                if user_answer and user_answer.strip():
                    score += 1
        
        # 현재 시간
        timestamp = data.get('startTime', datetime.now().isoformat())
        
        # 결과 저장
        conn = sqlite3.connect('quiz_results.db')
        c = conn.cursor()
        
        c.execute('''
            INSERT INTO results 
            (timestamp, user_id, name, center, service, role, score, total, 
             time_spent, timeout, questions_data, answers_data)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            timestamp,
            data.get('userId'),
            data.get('userName'),
            data.get('center'),
            data.get('service'),
            data.get('role'),
            score,
            total,
            data.get('timeSpent', 0),
            1 if data.get('isAutoSubmit') else 0,
            json.dumps(questions, ensure_ascii=False),
            json.dumps(answers, ensure_ascii=False)
        ))
        
        conn.commit()
        conn.close()
        
        # 채점 결과 반환
        return jsonify({
            'success': True,
            'totalScore': score,
            'correctCount': score,
            'total': total,
            'percentage': round((score / total * 100) if total > 0 else 0, 1)
        })
        
    except Exception as e:
        print(f"❌ 제출 오류: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/results')
def get_results():
    conn = sqlite3.connect('quiz_results.db')
    c = conn.cursor()
    
    # 필터링
    center = request.args.get('center')
    service = request.args.get('service')
    month = request.args.get('month')
    search = request.args.get('search')
    
    query = "SELECT * FROM results WHERE 1=1"
    params = []
    
    if center:
        query += " AND center = ?"
        params.append(center)
    if service:
        query += " AND service = ?"
        params.append(service)
    if month:
        query += " AND strftime('%Y-%m', created_at) = ?"
        params.append(month)
    if search:
        query += " AND (name LIKE ? OR user_id LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%'])
    
    query += " ORDER BY created_at DESC"
    
    c.execute(query, params)
    rows = c.fetchall()
    
    results = []
    for row in rows:
        results.append({
            'id': row[0],
            'timestamp': row[1],
            'userId': row[2],
            'name': row[3],
            'center': row[4],
            'service': row[5],
            'role': row[6],
            'score': row[7],
            'total': row[8],
            'timeSpent': row[9],
            'timeout': row[10],
            'questions': json.loads(row[11]) if row[11] else [],
            'answers': json.loads(row[12]) if row[12] else []
        })
    
    conn.close()
    return jsonify(results)

@app.route('/api/export-excel')
def export_excel():
    conn = sqlite3.connect('quiz_results.db')
    c = conn.cursor()
    
    c.execute("SELECT * FROM results ORDER BY created_at DESC")
    rows = c.fetchall()
    conn.close()
    
    # 엑셀 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "시험결과"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더
    headers = ['날짜', '이름', 'ID', '센터', '서비스', '역할', '점수', '총문제수', 
               '소요시간(초)', '시간초과', '문제', '선택값']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 데이터
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1).value = row[1]  # timestamp
        ws.cell(row=row_idx, column=2).value = row[3]  # name
        ws.cell(row=row_idx, column=3).value = row[2]  # user_id
        ws.cell(row=row_idx, column=4).value = row[4]  # center
        ws.cell(row=row_idx, column=5).value = row[5]  # service
        ws.cell(row=row_idx, column=6).value = row[6]  # role
        ws.cell(row=row_idx, column=7).value = row[7]  # score
        ws.cell(row=row_idx, column=8).value = row[8]  # total
        ws.cell(row=row_idx, column=9).value = row[9]  # time_spent
        ws.cell(row=row_idx, column=10).value = "예" if row[10] else "아니오"  # timeout
        
        # 문제 정보
        questions = json.loads(row[11]) if row[11] else []
        questions_str = "\n".join([f"Q{i+1}: {q['question']}" for i, q in enumerate(questions)])
        ws.cell(row=row_idx, column=11).value = questions_str
        
        # 선택값 정보
        answers = json.loads(row[12]) if row[12] else []
        answers_str = "\n".join([f"Q{i+1}: {a}" for i, a in enumerate(answers)])
        ws.cell(row=row_idx, column=12).value = answers_str
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 50
    ws.column_dimensions['L'].width = 50
    
    # 메모리에 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'quiz_results_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/api/admin/prompts', methods=['GET', 'POST'])
def admin_prompts():
    conn = sqlite3.connect('quiz_results.db')
    c = conn.cursor()
    
    if request.method == 'POST':
        data = request.json
        c.execute('''
            INSERT INTO admin_prompts 
            (month, service, custom_prompt, difficulty, created_by)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            data['month'],
            data['service'],
            data['customPrompt'],
            data['difficulty'],
            data.get('createdBy', 'admin')
        ))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    else:
        c.execute('''
            SELECT month, service, custom_prompt, difficulty, created_at 
            FROM admin_prompts 
            WHERE is_active = 1 
            ORDER BY created_at DESC
        ''')
        rows = c.fetchall()
        
        prompts = []
        for row in rows:
            prompts.append({
                'month': row[0],
                'service': row[1],
                'customPrompt': row[2],
                'difficulty': row[3],
                'createdAt': row[4]
            })
        
        conn.close()
        return jsonify(prompts)

@app.route('/health')
def health():
    return jsonify({
        'status': 'healthy',
        'loaded_services': list(SERVICE_POLICIES.keys()),
        'policy_count': len(SERVICE_POLICIES)
    })

if __name__ == '__main__':
    print("\n" + "="*60)
    print("📚 정책 파일 로딩 중...")
    print("="*60)
    
    load_policies()
    init_db()
    
    print("\n" + "="*60)
    print("🎯 KMCC 서비스 정책 테스트 시작!")
    print("="*60)
    print("📱 브라우저에서 접속: http://localhost:5001")
    print("📋 지원 서비스:", ', '.join(SERVICES))
    print("💾 데이터 저장: SQLite (quiz_results.db)")
    print("🛑 종료: Ctrl + C")
    print("="*60 + "\n")

     # Railway를 위한 포트 설정
    port = int(os.environ.get('PORT', 5001))
    app.run(debug=False, host='0.0.0.0', port=port)  # debug=False로 변경
